#!/usr/bin/env /usr/local/bin/pythonw
"""
INSTALLATION/REQUIREMENTS
PyAsset requires Python (>=3.7) and wxPython.

COPYRIGHT/LICENSING
Copyright (c) 2016-2025 Joseph J. Gorak. All rights reserved.
This code is in development -- use at your own risk. Email
comments, patches, complaints to joegorak808@outlook.com

This program is free software; you can redistribute it and/or
modify it under the terms of the GNU General Public License
as published by the Free Software Foundation; either version 2
of the License, or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program; if not, write to the Free Software
Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA  02111-1307, USA.
"""
#  Version information
#  06/11/2016     Initial version v0.1
#  08/07/2021     Version v0.2
#  04/02/2023     Version v1.0
#  01/01/2024     Version v1.1       JJG Rewrote ProcessTranactionSheet to handle comments and be more efficient

# TO-DOs
#

from ast import Continue
from math import e
from os import error
from openpyxl.reader.excel import load_workbook

from Asset import Asset
from AssetList import AssetList
from TransactionList import TransactionList
from Transaction import Transaction
from BillList import BillList
from Bill import Bill
from Date import Date

import wx

class ExcelToAsset(wx.Frame):
    def __init__(self, parent, title="ExcelToAsset", ignore_sheets=[]):
        self.parent = parent
        if parent != None:
             super(ExcelToAsset, self).__init__(parent, title=title)
        self.wb = None
        self.ignore_sheets = ignore_sheets

    def OpenXLSMFile(self, FileName):
        try:
            self.wb = load_workbook(FileName, read_only=True, data_only=True)
        except:
            error = "No such file or directory: " + FileName
            self.MsgBox(error)

    def ProcessAssetsSheet(self, parent, ignoredHeadings = ['Who']):
        self.parent = parent
        AssetsFound = AssetList(parent)
        if self.wb == None:
            return AssetsFound
        ws = self.wb["Assets"]
        for row in ws.rows:
            cv = row[0].value
            if cv == None or "Bills" in cv or "Total" in cv or "Cash Flow" in cv:
                continue
            elif "Accounts" in cv or "Other" in cv:
                ColumnHeaders = dict()
                col_num = 1
                for cell in row:
                    cv = cell.value
                    if cv != None:
                        if col_num == 1:
                            headerValue = "Name"
                        else:
                            headerValue = cv
                        ColumnHeaders[col_num] = headerValue
                        col_num += 1
                    else:
                        break
                    if len(ColumnHeaders) == 1:
                        continue
#               print(ColumnHeaders)
            else:
                col_num = 1
                for cell in row:
                    cv = cell.value
                    if col_num == 1:
                        # First column means this is a new asset... save name and pointer to the new asset location for later
                        # Also set type of asset using clues from the account name
                        new_asset = AssetsFound.get_asset_by_name(cv)
                        if new_asset == None:
                            new_asset = Asset(name = cv)
                            AssetsFound.append_by_object(new_asset)
                        asset_name = new_asset.get_name()
                        asset_name_upper = asset_name.upper()
                        if "HOUSE" in asset_name_upper:
                            new_asset.set_type("House")
                        elif "CAR" in asset_name_upper:
                            new_asset.set_type("Car")
                        elif "CHECKING" in asset_name_upper:
                            new_asset.set_type("Checking and Savings")
                        elif "SAVINGS" in asset_name_upper:
                            new_asset.set_type("Checking and Savings")
                        elif "MONEY MARKET" in asset_name_upper:
                            new_asset.set_type("Money Market")
                        elif "OVERDRAFT" in asset_name_upper:
                            new_asset.set_type("Overdraft")
                        elif "TSP" in asset_name_upper or "ANNUITY" in asset_name_upper or "LIFE" in asset_name_upper or "ROTH" in asset_name_upper or "IRA" in asset_name_upper:
                            new_asset.set_type("Retirement")
                        elif "DISCOVER" in asset_name_upper or "VISA" in asset_name_upper or "MC" in asset_name_upper or "MASTER CARD" in asset_name_upper or "BLUE" in asset_name_upper or "CREDIT CARD" in asset_name_upper:
                            new_asset.set_type("Credit Card")
                        elif "SEARS" in asset_name_upper or "MACY'S" in asset_name_upper:
                            new_asset.set_type("Store Card")
                        elif "LOAN" in asset_name_upper or "MORTGAGE" in asset_name_upper:
                            new_asset.set_type("Loan")
                        elif "EMERGENCY" in asset_name_upper:
                            new_asset.set_type("Emergency Fund")
                        else:
                            new_asset.set_type("Other")
                    else:  # 2nd and remaining columns are more data for the current asset...
                           # determine what field and update the object appropriately

                        heading = ColumnHeaders.get(col_num, "None")
                        if heading != "None":
                            if "Value (Curr)" in heading:
                                new_asset.set_value(cv)
                            elif "Value (Proj)" in heading:
                                new_asset.set_value_proj(cv)
                            elif "last pulled" in heading:
                                new_asset.set_last_pull_date(cv)
                            elif "Credit Limit" in heading:
                                new_asset.set_limit(cv)
                            elif "Avail (Online)" in heading:
                                new_asset.set_avail(cv)
                            elif "Avail (Proj)" in heading:
                                new_asset.set_avail_proj(cv)
                            elif "Estimate Method" in heading:
                                new_asset.set_est_method(cv)
                            elif "Rate" in heading:
                                new_asset.set_rate(cv)
                            elif "Payment" in heading:
                                new_asset.set_payment(cv)
                            elif "Due Date" in heading:
                                new_asset.set_due_date(cv)
                            elif "Sched" in heading:
                                new_asset.set_sched_date(cv)
                            elif "Min Pmt" in heading:
                                new_asset.set_min_pay(cv)
                            elif "Stmt Bal" in heading:
                                new_asset.set_stmt_bal(cv)
                            elif "Amt Over" in heading:
                                new_asset.set_amt_over(cv)
                            elif "Cash Limit" in heading:
                                new_asset.set_cash_limit(cv)
                            elif "Cash used" in heading:
                                new_asset.set_cash_used(cv)
                            elif "Cash avail" in heading:
                                new_asset.set_cash_avail(cv)
                            else:
                                found = False
                                for ignore in ignoredHeadings:
                                    if ignore in heading:
                                        found = True
                                if not found:
                                    print("ProcessAssetSheets: Unknown field " + heading + " ignored!")
                    col_num += 1
                    if col_num > len(ColumnHeaders):
                        break
        # if self.parent == None, this is the test code and we didn't set up the menus so don't change the menu choices
        test_parent = self.parent
        if AssetsFound != None and test_parent != None:
            self.Parent.fileMenuItem["Open"].Enable(False)
            self.Parent.fileMenuItem["ImportCSV"].Enable(False)
            self.Parent.fileMenuItem["ImportXLSX"].Enable(False)
            self.Parent.fileMenuItem["Save"].Enable(False)              # JJG 08/04/2024 Prevent accidentally overwriting .xlsx if they forget to change filename!
            self.Parent.fileMenuItem["SaveAs"].Enable(True)
            self.Parent.fileMenuItem["Close"].Enable(True)
            self.Parent.fileMenuItem["Export"].Enable(True)
            self.Parent.fileMenuItem["Archive"].Enable(True)
        return AssetsFound

    def GetTransactionSheetNames(self):
        return_sheets = []
        if self.wb == None:
            return return_sheets
        sheets = list(self.wb.sheetnames)
        for sheet in sheets:
            if sheet in self.ignore_sheets:
                continue
            else:
                return_sheets.append(sheet)
        return return_sheets

    def ProcessTransactionSheet(self, whichAsset, SheetName):
        TransactionsFound = TransactionList(whichAsset)
        if self.wb == None:
            return TransactionsFound
        ws = self.wb[SheetName]
        SheetName = str(SheetName).upper()
        if "CHECKING" in SheetName:
            ColumnHeaders = ["Pmt Method", "Chk #", "Payee", "Amt", "B", "Sched date", "Due date", "Comment"]
        else:
            ColumnHeaders = ["Pmt Method", "Payee", "Amt", "B", "Sched date", "Due date", "Comment"]
        placeIndex = {}
        row_num = -1
        for row in ws.rows:
            row_num += 1
            if row_num == 0:
                for index in range(len(ColumnHeaders)):
                    foundIndex = -1
                    for cell in row:
                        if cell.data_type != 's': continue
                        if ColumnHeaders[index].upper() == cell.value.upper():
                            foundIndex = cell.column-1
                            break
                    if foundIndex != -1:
                        placeIndex[ColumnHeaders[index].upper()] = foundIndex
                    else:
                        placeIndex[ColumnHeaders[index].upper()] = -1
            else:
                cv = row[0].value                           # JJG 1/1/2024 Don't waste time on rows without headers
                if cv == None:
                    continue
                else:
                    new_transaction = Transaction(self)
                    for heading in ColumnHeaders:
                        headingUpper = heading.upper()
                        if placeIndex[headingUpper] == -1: continue
                        cv =  row[placeIndex[headingUpper]].value
                        if cv == None: continue
                        if "PMT METHOD" == headingUpper:
                            new_transaction.set_pmt_method(cv)
                        elif "CHK #" == headingUpper:
                            pmt_method = new_transaction.get_pmt_method()
                            if pmt_method != None:
                                new_transaction.set_check_num(cv)
                        elif "PAYEE" == headingUpper:
                            payee = cv
                            new_transaction.set_payee(payee)
                            if payee[0:len("Paydown")]=="Paydown" or payee[0:len("Xfer")]=="Xfer":
                                new_transaction.set_action("+")
                            else:
                                new_transaction.set_action("-")
                        elif "AMT" == headingUpper:
                            new_transaction.set_amount(cv)
                        elif "B" == headingUpper:
                            if cv != "":
                                new_transaction.set_action(cv)
                        elif "SCHED DATE" == headingUpper:
                            if cv != "":
                                new_transaction.set_sched_date(cv)
                        elif "DUE DATE" == headingUpper:
                            if cv != "":
                                new_transaction.set_due_date(cv)
                        elif "COMMENT" == headingUpper:
                            if cv != "" and cv != None:
                                new_transaction.set_comment(cv)
                        else:
                           print("ProcessTransactionSheet: Unknown field " + headingUpper + " on sheet " + SheetName + " ignored!")

                    # At this point we have filled in all the necessaty fields so just make sure the new transaction hasa a scheduled date
                    # Don't add transactions that don't have scheduled date
                    if new_transaction.get_sched_date() == None or new_transaction.get_sched_date() == '':
                        continue

                    new_transaction.parent = self.parent                # make sure transaction gets attached to asset and not EXCEL object!  JJG 1/15/2023

                    # JJG 7/15/2025 Determine transaction state since it is not part of spreadsheet
                    if new_transaction.get_sched_date() != None and new_transaction.get_pmt_method() != "TBD":
                        new_transaction.set_state("scheduled")
                    if new_transaction.get_pmt_method() == "processing":
                        new_transaction.set_state("pending")
                    if new_transaction.get_pmt_method() == "TBD" and new_transaction.get_state() == "unknown" and new_transaction.get_amount() != 0.0:
                        new_transaction.set_comment("Need to schedule this ASAP!")
                    TransactionsFound.insert(new_transaction)
        return TransactionsFound

    def ProcessBillsSheet(self, haveParent=True):
        MAX_COLS_TO_PROCESS = 8             # Only these columns have data we want! JJG 4/1/2023
        BillsFound = BillList()
        BillPlaces = dict()
        if self.wb == None:
            return BillsFound
        ws = self.wb["Bills"]

        row_num = 0
        cur_type = "unknown"
        if haveParent:
            self.parent = wx.GetTopLevelParent(self).Parent
            title = self.GetTitle()
            super(ExcelToAsset, self).__init__(self.parent, title=title)
            new_bill = Bill(self.parent)
        else:
            new_bill = Bill(None)
        wb = None

        # Implement a state machine to process the Bill sheet!           JJG 9/25/2025

        ExcelReadStates = ["SectionHeader", "HeaderRow", "DataRow"]             # These are the states we can be in reading the excel file
        ExcelReadState = "SectionHeader"                                        # Start out expecting a section header
        possible_section_header = True
        Finished = False

        for row in list(ws.rows):
            if Finished:                                                        # Check for early termination of the state machine JJG 9/23/2025
                break
            row_num += 1
            if row_num == 1:                                                   # Ignore the first row which is just the title
                continue
            NextRow = False
            for col_num in range(0, MAX_COLS_TO_PROCESS):
                # If we set the flag while processing the current row, break out to go get the next row
                if NextRow:
                    break
                cv = row[col_num].value
                cv_orig = cv                                                   # Save the original value for later

                NextRow = False

                # If we are in the first column and not proessing field header rows, we might have a section header
                if col_num == 0 and ExcelReadState != "HeaderRow":
                    possible_section_header = True
                    check_col = 0
                    # It this row might be a section header, check the rest of the columns
                    # If we find anything but nothing in any column but this first, this isn't a section header so reset the flag
                    while possible_section_header and check_col != MAX_COLS_TO_PROCESS:
                        check_col += 1
                        test_cv = row[check_col].value
                        if test_cv != None:
                            possible_section_header = False

                    # Set ExcelReadState correctly based on what we found
                    if possible_section_header:
                        ExcelReadState = "SectionHeader"
                    else:
                        ExcelReadState = "DataRow"

                # If the column we just read in has a string value, convert it to all lower case to simply the rest of our processing
                if type(cv) is str:
                    cv = cv.lower()

                # if this is a section header line, record the type and go get the next row (which should be the header row)
                if ExcelReadState == "SectionHeader":
                    # For now, we won't process the TOTALS section                   TODO: See if we can create something useful later!   JJG 9/27/2025
                    if cv == "totals":
                        Finished = True
                        break

                    # Bill types in this code are "checking and savings", "credit card", "emergency fund" "loan", "expense" and "unknown
                    # But the section headers in the EXCEL file are "checking and savings accounts", "credit cards", "emergeny fund", "loans", and "expenses"
                    # so we extract the appropriate number of characters from the start of the string and see if we recognize it
                    # If we don't recognize it, set the type to "unknown" and use that as the bill type

                    bill_type_index = -1
                    if cv != None and col_num == 0:
                        bill_types = ["checking and savings", "credit card", "emergency fund", "loan", "expense", "unknown"]
                        for test_type in bill_types:
                            bill_type = cv[:len(test_type)]
                            try:
                                bill_type_index = bill_types.index(bill_type)
                                break
                            except:
                                bill_type_index = bill_types.index("unknown")

                        if bill_type_index != -1:
                            self.cur_type = bill_types[bill_type_index]
                        else:
                            self.cur_type = "unknown"

                    # Now that we have processed the section header, the next row should be the header row
                    new_bill.set_type(self.cur_type)
                    ExcelReadState = "HeaderRow"
                    NextRow = True
                    continue

                elif ExcelReadState == "HeaderRow":
                        BillPlaces[col_num] = cv
                        # if we just processed the last column that we need on this row, set NextRow flag and break out of col loop
                        # and the next row should be the start of the bill data so change the ReadState
                        if col_num == MAX_COLS_TO_PROCESS-1:
                            NextRow = True
                            ExcelReadState = "DataRow"
                            break

                elif ExcelReadState == "DataRow":

                    heading = BillPlaces.get(col_num, "None")   # Don't think we should ever get "None" here but just in case...

                    if heading != None:
                        if heading == "payee":
                            new_bill.set_payee(cv_orig)
                        elif heading == "amount":
                            new_bill.set_amount(cv_orig)
                        elif heading == "min due":
                            new_bill.set_min_due(cv_orig)
                        elif heading == "due date":
                            new_bill.set_due_date(cv_orig)
                        elif heading == "sched date":
                            new_bill.set_sched_date(cv_orig)
                        elif heading == "pmt acct":
                            new_bill.set_pmt_acct(cv_orig)
                        elif heading == "pmt method":
                            new_bill.set_pmt_method(cv_orig)
                        elif heading == "frequency":
                            new_bill.set_pmt_frequency(cv_orig)
                    else:
                        error = "ProcessBillsSheet: Unknown field " + str(heading) + " on row " + str(row_num) + " ignored!"
                        self.MsgBox(error)
                else:
                    error = "Internal error in ExcelToAsset.ProcessBillsSheet: Unexpected ExcelReadState "
                    error += ExcelReadState + "\nValid ExecReadStates are " + str(ExcelReadStates)
                    self.MsgBox(error)
                    return None
            
            # At this point we found a section we don't want to process yet ==> set Finish
            # or we finished processing all the rows in a HeaderRow ==> set NextRow to go get the next row
            # If we set the NextRow flag, break out of the column processing loop and go get the next row                       

            if Finished:
                break
            if NextRow:
               continue

#            print("Inserting bill ", new_bill)         # Uncomment this to see the bills as they are inserted (it also helped debug the code!))
            BillsFound.insert(new_bill)
            if self.parent != None:                     # Get a new new_bill structure for the next row and if we are the test code, don't set the parent
                new_bill = Bill(self.parent)
            else:
                new_bill = Bill(None)
            new_bill.set_type(self.cur_type)
 
        # At this point bills were inserted in the order they were found in the Bill sheet.
        # Now do a multi-level sort on the list of bills.  JJG 1/25/2025
        BillsFound.sort_by_fields(BillList.getSortOrder(self))
        return BillsFound

    def MsgBox(self, message):
        if self.parent != None:
            d = wx.MessageDialog(self, message, "error", wx.OK | wx.ICON_INFORMATION)
            d.ShowModal()
            d.Destroy()
        else:
            d = wx.MessageDialog(None, message, "error", wx.OK | wx.ICON_INFORMATION)
            d.ShowModal()
            d.Destroy()


if __name__ == '__main__':
    import sys
    import wx
    from AssetFrame import AssetFrame

    class testFrame(wx.Frame):
        test_filename = 'C:\\Users\\joego\\Desktop\\Finances for budgeting.xlsm';
        ignore_sheets = ['Assets', 'Bills'];

        def __init__(self):
            super().__init__(parent=None, title="ExceltoAsset tester")
            self.assetFrame = AssetFrame(self, 'PyAsset', "", "")
            self.assetFrame.Show(False)                             # We don't need (or want!) to see the Asset Frame JJG 9/22/2025
            self.Bind(wx.EVT_CLOSE,self.close)
            
            panel = wx.Panel(self)

            # The widget that will receive the keystrokes
            self.text_ctrl = wx.TextCtrl(panel, style=wx.TE_MULTILINE)
        
            # A button to trigger the testing
            button = wx.Button(panel, label='do tests')
            button.Bind(wx.EVT_BUTTON, self.doTests)

            # Layout
            sizer = wx.BoxSizer(wx.VERTICAL)
            sizer.Add(self.text_ctrl, 1, wx.EXPAND | wx.ALL, 5)
            sizer.Add(button, 0, wx.ALIGN_CENTER | wx.ALL, 5)
            panel.SetSizer(sizer)

            self.Show()


        def testAssets(self):
            etoa = ExcelToAsset(None, self.ignore_sheets)
            etoa.OpenXLSMFile(self.test_filename)
            AssetsFound = etoa.ProcessAssetsSheet(None)
            if AssetsFound != None:
                print("Found " + str(len(AssetsFound)) + " assets!")
                for asset in AssetsFound:
                    print(asset)
            self.AssetList = AssetsFound

        def testTransactions(self):
            etoa = ExcelToAsset(None, self.ignore_sheets)
            etoa.OpenXLSMFile(self.test_filename)
            SheetsFound = etoa.GetTransactionSheetNames()
            print("Found " + str(len(SheetsFound)) + " transaction sheets!")
            for sheet in SheetsFound:
                if sheet in self.ignore_sheets:
                    print("Ignoring sheet ", sheet)
                    continue
                print("Processing sheet " + sheet)
                TransactionsFound = etoa.ProcessTransactionSheet(self.AssetList.get_asset_by_name(sheet), sheet)
                if TransactionsFound != None:
                    print("Found " + str(len(TransactionsFound)) + " transactions on sheet " + sheet + "!")
                    if len(TransactionsFound) != 0:
                        for transaction in TransactionsFound:
                            print(transaction)
    
        def testBills(self):
            etoa = ExcelToAsset(None, self.ignore_sheets)
            etoa.OpenXLSMFile(self.test_filename)
            BillsFound = etoa.ProcessBillsSheet(False)
            if BillsFound != None:
                print("Found " + str(len(BillsFound)) + " bills!")
                if BillsFound.getBills() != None:
                    for bill in BillsFound:
                        print(bill)

        def doTests(self, event):
            print("Testing loading assets\n\n")
            self.testAssets()
            print("\nTesting loading transactions\n\n")
            self.testTransactions()
            print("\nTesting loading bills\n\n")
            self.testBills()
            print("\nExcelToAsset testing done!")

        def close(self, event):
            self.assetFrame.Destroy()
            self.Destroy()

    app = wx.App(False)  # Create a new app, don't redirect stdout/stderr to a window.
    methods = [method for method in dir(ExcelToAsset) if callable(getattr(ExcelToAsset, method)) and not method.startswith("__")]
    parent_methods = [method for method in dir(wx.Frame) if callable(getattr(wx.Frame, method)) and not method.startswith("__")]
    new_methods = []
    for m in methods:
        if not m in parent_methods:
            new_methods.append(m)
    print("Methods added by ExcelToAsset:"+str(new_methods))
    app.frame = testFrame()
    app.frame.Show()
    app.MainLoop()
